import sys
import configparser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout,
    QWidget,  QScrollArea,
    QTabWidget, QGroupBox, QGridLayout, QMessageBox, QLabel, QLineEdit, QDialog, QFormLayout, QFileDialog, QSpinBox,
    QCheckBox, QComboBox, QSizePolicy, QDialogButtonBox, QTextEdit
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIntValidator
import json

from globalFunctions import get_float_value, get_int_value

# ComponentClass
import componentClass
import searchComponentInShopClass

# Config file name
CONFIG_FILE = "config.ini"


# Configuration window
class ConfigurationWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Config")
        self.setFixedSize(400, 300)

        #Get the current configuration. If it not exits, create it
        self.config = self.getConfigurationFromIniFile()

        # Crear el widget de pestañas
        self.tabs = QTabWidget()

        # Mouser TAB
        self.mouser_tab = QWidget()
        self.mouser_layout = QFormLayout()
        self.mouser_token = QLineEdit()
        self.mouser_checkbox = QCheckBox("Enable search on Mouser")
        self.mouser_currency = QComboBox()
        self.mouser_currency.addItems(["EUR", "USD"])
        self.mouser_layout.addRow("Mouser Token:", self.mouser_token)
        self.mouser_layout.addRow(self.mouser_checkbox)
        self.mouser_tab.setLayout(self.mouser_layout)

        # DigiKey TAB
        self.digikey_tab = QWidget()
        self.digikey_layout = QFormLayout()
        self.digikey_token = QLineEdit()
        self.digikey_secret = QLineEdit()
        self.digikey_checkbox = QCheckBox("Enable search on DigiKey")
        self.digikey_local_site = QComboBox()
        self.digikey_local_site.addItems([
            "AT", "AU", "BE", "BG", "BR", "CA", "CH", "CN", "CZ", "DE", "DK", "EE", "ES", "FI", "FR", "GR", "HK",
            "HU", "IE", "IL", "IN", "IT", "JP", "KR", "LT", "LU", "LV", "MX", "MY", "NL", "NO", "NZ", "PH", "PL", "PT",
            "RO", "SE", "SG", "SI", "SK", "TH", "TW", "UK", "US", "ZA"
        ])

        self.digikey_language = QComboBox()
        self.digikey_language.addItems([
            "CS", "DA", "DE", "EN", "ES", "FI", "FR", "HE", "HU", "IT", "JA", "KO", "NL", "NO", "PL", "PT", "RO",
            "SV", "TH", "ZHS", "ZHT"
        ])
        self.digikey_currency = QComboBox()
        self.digikey_currency.addItems([
            "AUD", "CAD", "CHF", "CNY", "CZK", "DKK", "EUR", "GBP", "HKD", "HUF", "ILS", "INR", "JPY", "KRW", "MYR",
            "NOK", "NZD", "PHP", "PLN", "RON", "SEK", "SGD", "THB", "TWD", "USD", "ZAR"
        ])
        self.digikey_layout.addRow("DigiKey ClientID:", self.digikey_token)
        self.digikey_layout.addRow("DigiKey ClientSecret:", self.digikey_secret)
        self.digikey_layout.addRow("Currency:", self.digikey_currency)
        self.digikey_layout.addRow("LocalSite:", self.digikey_local_site)
        self.digikey_layout.addRow("Language:", self.digikey_language)
        self.digikey_layout.addRow(self.digikey_checkbox)
        self.digikey_tab.setLayout(self.digikey_layout)

        # TME TAB
        self.tme_tab = QWidget()
        self.tme_layout = QFormLayout()
        self.tme_token = QLineEdit()
        self.tme_secret = QLineEdit()
        self.tme_checkbox = QCheckBox("Enable search on TME")

        self.tme_local_site = QComboBox()
        self.tme_local_site.addItems([
            "PL", "ES", "US", "GB","FR",
        ])

        self.tme_language = QComboBox()
        self.tme_language.addItems([
            "en", "pl", "de",
        ])

        self.tme_layout.addRow("TME Token:", self.tme_token)
        self.tme_layout.addRow("TME Secret:", self.tme_secret)
        self.tme_layout.addRow("Language:", self.tme_language)
        self.tme_layout.addRow("Country:", self.tme_local_site)

        self.tme_layout.addRow(self.tme_checkbox)
        self.tme_tab.setLayout(self.tme_layout)

        # Element14 TAB
        self.element14_tab = QWidget()
        self.element14_layout = QFormLayout()
        self.element14_token = QLineEdit()
        self.element14_checkbox = QCheckBox("Enable search on Element14")
        self.element14_market = QComboBox()
        self.element14_market.addItems([
            "bg.farnell.com", "cz.farnell.com", "dk.farnell.com", "at.farnell.com",
            "ch.farnell.com", "de.farnell.com", "cpc.farnell.com", "cpcireland.farnell.com",
            "export.farnell.com", "onecall.farnell.com", "ie.farnell.com", "il.farnell.com",
            "uk.farnell.com", "es.farnell.com", "ee.farnell.com", "fi.farnell.com",
            "fr.farnell.com", "hu.farnell.com", "it.farnell.com", "lt.farnell.com",
            "lv.farnell.com", "be.farnell.com", "nl.farnell.com", "no.farnell.com",
            "pl.farnell.com", "pt.farnell.com", "ro.farnell.com", "ru.farnell.com",
            "sk.farnell.com", "si.farnell.com", "se.farnell.com", "tr.farnell.com",
            "canada.newark.com", "mexico.newark.com", "www.newark.com", "cn.element14.com",
            "au.element14.com", "nz.element14.com", "hk.element14.com", "sg.element14.com",
            "my.element14.com", "ph.element14.com", "th.element14.com", "in.element14.com",
            "tw.element14.com", "kr.element14.com", "vn.element14.com"
        ])

        self.element14_layout.addRow("Element14 Token:", self.element14_token)
        self.element14_layout.addRow("Markets:", self.element14_market)
        self.element14_layout.addRow(self.element14_checkbox)
        self.element14_tab.setLayout(self.element14_layout)


        # Add tab to widget
        self.tabs.addTab(self.mouser_tab, "Mouser")
        self.tabs.addTab(self.digikey_tab, "DigiKey")
        self.tabs.addTab(self.tme_tab, "TME")
        self.tabs.addTab(self.element14_tab, "Element14")

        # Save button
        self.save_button = QPushButton("Guardar")
        self.save_button.clicked.connect(self.save_settings)

        # Settings layour
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tabs)
        main_layout.addWidget(self.save_button)
        self.setLayout(main_layout)

        # Load settings to settings view
        self.load_settings()

    # Function to get config from ini file. If not exists, create it
    def getConfigurationFromIniFile(self):
        config = configparser.ConfigParser()
        if not config.read(CONFIG_FILE):
            config["TOKENS"] = {}
            config["TOKENS_SECRET"] = {}
            config["ENABLED"]    = {"mouser": "False", "digikey": "False", "tme": "False", "element14": "False"}
            config["MARKET"]    = {"mouser_market": "None", "digikey_market": "None", "element14_market": "None"}
            config["CURRENCY"]   = {"mouser_currency": "EUR", "digikey_currency": "EUR", "element14_currency": "EUR"}
            config["LANGUAGE"]   = {"mouser_language": "", "digikey_language": "ES ", "tme_language": "EUR"}
            config["LOCAL_SITE"] = {"mouser_local_site": "EUR", "digikey_local_site": "EUR", "tme_local_site": "EUR"}
            with open(CONFIG_FILE, "w") as configfile:
                config.write(configfile)
        return config

    #Load settings to config view
    def load_settings(self):
        # Mouser settings
        self.mouser_token.setText(self.config.get("TOKENS", "mouser_token", fallback=""))
        self.mouser_checkbox.setChecked(self.config.getboolean("ENABLED", "mouser", fallback=False))
        self.mouser_currency.setCurrentText(self.config.get("CURRENCY", "mouser_currency", fallback="EUR"))

        # Digikey settings
        self.digikey_token.setText(self.config.get("TOKENS", "digikey_token", fallback=""))
        self.digikey_secret.setText(self.config.get("TOKENS_SECRET", "digikey_token_secret", fallback=""))
        self.digikey_checkbox.setChecked(self.config.getboolean("ENABLED", "digikey", fallback=False))
        self.digikey_currency.setCurrentText(self.config.get("CURRENCY", "digikey_currency", fallback="EUR"))
        self.digikey_local_site.setCurrentText(self.config.get("LOCAL_SITE", "digikey_local_site", fallback="EUR"))
        self.digikey_language.setCurrentText(self.config.get("LANGUAGE", "digikey_language", fallback="EUR"))

        # TME settings
        self.tme_token.setText(self.config.get("TOKENS", "tme_token", fallback=""))
        self.tme_secret.setText(self.config.get("TOKENS_SECRET", "tme_token_secret", fallback=""))
        self.tme_checkbox.setChecked(self.config.getboolean("ENABLED", "tme", fallback=False))
        self.tme_local_site.setCurrentText(self.config.get("LOCAL_SITE", "tme_local_site", fallback="EUR"))
        self.tme_language.setCurrentText(self.config.get("LANGUAGE", "tme_language", fallback="EUR"))

        # Element14_token settings
        self.element14_token.setText(self.config.get("TOKENS", "element14_token", fallback=""))
        self.element14_checkbox.setChecked(self.config.getboolean("ENABLED", "element14", fallback=False))
        self.element14_market.setCurrentText(self.config.get("MARKETS", "element14_market", fallback="es.farnell.com"))

    # Save settings to .ini file
    def save_settings(self):
        # Mouser settings
        self.config.set("TOKENS", "mouser_token", self.mouser_token.text())
        self.config.set("ENABLED", "mouser", str(self.mouser_checkbox.isChecked()))
        self.config.set("CURRENCY", "mouser_currency", self.mouser_currency.currentText())

        # Digikey settings
        self.config.set("TOKENS", "digikey_token", self.digikey_token.text())
        self.config.set("TOKENS_SECRET", "digikey_token_secret", self.digikey_secret.text())
        self.config.set("ENABLED", "digikey", str(self.digikey_checkbox.isChecked()))
        self.config.set("CURRENCY", "digikey_currency", self.digikey_currency.currentText())
        self.config.set("LANGUAGE", "digikey_language", self.digikey_language.currentText())
        self.config.set("LOCAL_SITE", "digikey_local_site", self.digikey_local_site.currentText())

        # TME settings
        self.config.set("TOKENS", "tme_token", self.tme_token.text())
        self.config.set("TOKENS_SECRET", "tme_token_secret", self.tme_secret.text())
        self.config.set("ENABLED", "tme", str(self.tme_checkbox.isChecked()))
        self.config.set("LANGUAGE", "tme_language", self.tme_language.currentText())
        self.config.set("LOCAL_SITE", "tme_local_site", self.tme_local_site.currentText())

        # Element14 settings
        self.config.set("TOKENS", "element14_token", self.element14_token.text())
        self.config.set("ENABLED", "element14", str(self.element14_checkbox.isChecked()))
        self.config.set("MARKETS", "element14_market", self.element14_market.currentText())

        # Escribir archivo de configuración
        with open(CONFIG_FILE, "w") as configfile:
            self.config.write(configfile)

        self.accept()


# Main window
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Load configuration settings
        self.config = self.load_config()

        # Set window appearance and styles
        self.set_window_style()
        # apply_stylesheet(app, theme='light_purple.xml')

        # Set the window title and dimensions
        self.setWindowTitle("IntelligentElectronicComponentSearch")
        self.setGeometry(100, 100, 1200, 600)

        # List containing all components as ComponentClass objects
        self.componentsClassList = []  # List of ComponentClass instances

        # Layout for storing component widgets
        self.componentsWidgetsLayout = []

        # Central widget for the main window
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout for the application
        main_layout = QVBoxLayout()

        # Top button bar
        button_layout = QHBoxLayout()
        buttons = [
            ("NEW PROJECT", self.clickButtonAction_newProject),
            ("OPEN PROJECT", self.clickButtonAction_openProject),
            ("SAVE PROJECT", self.clickButtonAction_saveProject),
            ("SEARCH COMPONENT IN SHOPS", self.clickButtonAction_searchComponentsInShops),
            ("GENERATE OPTIMIZED BOM", self.clickButtonAction_generateOptimizedBom),
            ("CONFIGURATION", self.clickButtonAction_openConfiguration)
        ]
        # Create buttons and connect them to their respective callbacks
        for text, callback in buttons:
            button = QPushButton(text)
            button.clicked.connect(callback)
            button_layout.addWidget(button)

        # Lower configuration section
        config_layout = QHBoxLayout()

        # Board number configuration
        self.board_number_label = QLabel("BOARDS NUMBER:")
        self.board_number_spinbox = QSpinBox()
        self.board_number_spinbox.setMinimum(1)
        self.board_number_spinbox.setMaximum(10000)
        self.board_number_spinbox.setValue(1)
        self.board_number_spinbox.valueChanged.connect(self.componentView_UpdateTotalQuantitiesField)

        # Number of components configuration
        self.components_label = QLabel("NUMBER OF COMPONENTS:")
        self.components_spinbox = QSpinBox()
        self.components_spinbox.setMinimum(0)
        self.components_spinbox.setValue(len(self.componentsWidgetsLayout))
        self.components_spinbox.valueChanged.connect(self.componentView_UpdateComponentList)

        # Add widgets to the configuration layout
        config_layout.addWidget(self.board_number_label)
        config_layout.addWidget(self.board_number_spinbox)
        config_layout.addWidget(self.components_label)
        config_layout.addWidget(self.components_spinbox)

        # Scrollable area container
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        # Container for component widgets
        self.components_widget = QWidget()
        self.components_layout = QVBoxLayout()
        self.components_widget.setLayout(self.components_layout)

        # Add the component widget to the scroll area
        self.scroll_area.setWidget(self.components_widget)

        # Add layouts to the main layout
        main_layout.addLayout(button_layout)
        main_layout.addLayout(config_layout)
        main_layout.addWidget(self.scroll_area)

        # Set the central widget's layout
        central_widget.setLayout(main_layout)

        # Initialize the component view with the number of widgets
        self.initialize_components(len(self.componentsWidgetsLayout))

    # Set style
    def set_window_style(self):
        self.setStyleSheet("""
            QGroupBox {
                font: bold 12px;
                border: 2px solid #5dade2;
                border-radius: 10px;
                margin-top: 10px;
                background-color: #d6eaf8;
            }
            QPushButton {
            background-color: #5dade2;
            border: 2px solid #2980b9;
            border-radius: 8px;
            color: #1d334a;
            padding: 10px;
            font: bold 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
                color: #2471a3;
                background-color: #aed6f1;
                border-radius: 5px;
            }

            QLabel {
               font: 13px Arial, sans-serif;
                color: #2C3E50; 
            }
            

            QLineEdit, QSpinBox {
                border: 1px solid #5dade2;
                border-radius: 5px;
                padding: 5px;
                background-color: #ebf5fb;
                color: #1b4f72;
            }

            QLineEdit:focus, QSpinBox:focus {
                border: 2px solid #2980b9;
                background-color: #ffffff;
            }

            QLabel:hover {
                color: #2980b9;
            }
        """)


    # Create a new componentView with their style
    def componenteView_createIndividualComponentView(self, component: componentClass.ComponentClass):
        component_group = QGroupBox(f"Component {len(self.componentsWidgetsLayout)}")
        component_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        #component_group.setFixedSize(1100, 150)
        component_group.setStyleSheet(
            "QGroupBox {"
            "    font: bold 12px;"
           # "    border: 2px solid #5dade2;"
            "    border-radius: 10px;"
            "    margin-top: 10px;"
            #"    background-color: #d6eaf8;"
            "}"
            "QGroupBox::title {"
            "    subcontrol-origin: margin;"
            "    subcontrol-position: top left;"
            "    padding: 0 5px;"
            #"    color: #2471a3;"
            #"    background-color: #aed6f1;"
            "    border-radius: 5px;"
            "}"
            "QLabel {"
            "    font: 11px;"
            #"    color: #154360;"
            "}"
            "QLineEdit, QSpinBox {"
           # "    border: 1px solid #5dade2;"
            "    border-radius: 5px;"
            "    padding: 5px;"
            #"    background-color: #ebf5fb;"
            #"    color: #1b4f72;"
            "}"
            "QLineEdit:focus, QSpinBox:focus {"
           # "    border: 2px solid #2980b9;"
            #"    background-color: #ffffff;"
            "}"
            #"QLabel:hover {"
            #"    color: #2980b9;"
            #"}"
        )

        component_layout = QGridLayout()

        # Primera columna
        first_column = QGridLayout()
        first_column.setAlignment(Qt.AlignTop)
        first_column.addWidget(QLabel("Part Number:"), 0, 0)
        part_number_edit = QLineEdit()
        part_number_edit.setObjectName("OBJECT_NAME_PART_NUMBER")
        part_number_edit.setText(component.part_number);
        first_column.addWidget(part_number_edit, 0, 1)

        first_column.addWidget(QLabel("Manufacturer:"), 1, 0)
        manufacturer_edit = QLineEdit()
        manufacturer_edit.setObjectName("OBJECT_NAME_MANUFACTURER")
        manufacturer_edit.setText(component.manufacturer);
        first_column.addWidget(manufacturer_edit, 1, 1)

        first_column.addWidget(QLabel("Units:"), 2, 0)
        units_edit = QLineEdit()
        units_edit.setObjectName("OBJECT_NAME_UNIT")
        units_edit.setValidator(QIntValidator(0, 100000))
        units_edit.setText(str(component.units_per_board));
        first_column.addWidget(units_edit, 2, 1)

        first_column.addWidget(QLabel("Total Units:"), 3, 1)
        total_units_label = QLabel("0")
        total_units_label.setObjectName("OBJECT_NAME_TOTAL_UNITS")
        total_units_label.setText(str(component.total_units_per_board));
        first_column.addWidget(total_units_label, 3, 2)

        units_edit.textChanged.connect(
            lambda value, lbl=total_units_label: lbl.setText(value if value.isdigit() else "0")
        )

        # Tercera columna (7/10)
        third_column = QGridLayout()
        third_column.setAlignment(Qt.AlignTop)
        third_column.addWidget(QLabel("MARKET"), 0, 0)

        for index, label_text in enumerate(componentClass.attributes):
            if label_text != "Market":  # Skip "Market"
                if componentClass.attributes_visible[index] == 0:
                    continue
                third_column.addWidget(QLabel(label_text), 0, index + 1)  # Adjust the column index accordingly

        for row, market in enumerate(componentClass.markets, start=1):
            # Recupera información del market
            market_info = component.get_market_info(market)

            third_column.addWidget(QLabel(market), row, 0)
            for idx, attribute in enumerate(componentClass.attributes):
                if componentClass.attributes_visible[idx] == 0:
                    continue

                # Crear un QLineEdit
                line_edit = QLineEdit()

                background_color = "#ebf5fb"

                if 'INFO_ALERT' in market_info:
                    if (market_info['INFO_ALERT'] == 1):
                        background_color = "#efedd4"
                    elif (market_info['INFO_ALERT'] == 2):
                        background_color = "#efd4e4"
                #else:
                    #background_color = "#d4efdf"

                line_edit.setStyleSheet(
                    f"QLineEdit {{"
                    f"    border: 0px solid #82e0aa;"  # Borde verde
                    f"    border-radius: 5px;"
                    f"    background-color: {background_color};"  # Usando la variable de fondo
                    f"}}"
                    f"QLineEdit:focus {{"
                    f"    border: 0px solid #27ae60;"  # Borde verde más oscuro cuando está enfocado
                    f"    background-color: #ffffff;"  # Fondo blanco cuando está enfocado
                    f"}}"
                )

                # Generar un nombre único combinando market y attribute
                object_name = f"{market}_{attribute.replace(' ', '_')}"
                line_edit.setObjectName(object_name)
                line_edit.setText(market_info[attribute])

                # Agregar el QLineEdit al layout
                third_column.addWidget(line_edit, row, idx + 1)

        # Combinar las dos columnas
        component_layout.addLayout(first_column, 0, 0, 1, 3)  # Ocupa 3/10
        component_layout.addLayout(third_column, 0, 3, 1, 7)  # Ocupa 7/10

        component_layout.addLayout(first_column, 0, 0, 1, 3)
        component_group.setLayout(component_layout)
        return component_group

    def load_config(self):
        config = configparser.ConfigParser()
        if not config.read(CONFIG_FILE):
            config["TOKENS"] = {}
            config["TOKENS_SECRET"] = {}
            config["ENABLED"] = {"mouser": "False", "digikey": "False", "tme": "False", "element14": "False"}
            config["CURRENCY"] = {"mouser_currency": "EUR", "digikey_currency": "EUR", "tme_currency": "EUR"}
            config["LANGUAGE"] = {"digikey_language": "ES", "tme_language": "es"}
            config["LOCAL_SITE"] = {"digikey_local_site": "EUR", "tme_local_site": "EUR"}
            config["MARKETS"] = {}

            with open(CONFIG_FILE, "w") as configfile:
                config.write(configfile)
        return config


    # Initilice n components
    def initialize_components(self, count):
        for _ in range(count):
            component = self.componenteView_createIndividualComponentView()
            self.componentView_AddComponent(component)

    # Add component to view
    def componentView_AddComponent(self, component):
        self.components_layout.addWidget(component)
        self.componentsWidgetsLayout.append(component)

    # Remove component view
    def componentView_RemoveAllComponents(self):
        while self.componentsWidgetsLayout:
            component = self.componentsWidgetsLayout.pop()
            self.components_layout.removeWidget(component)
            component.deleteLater()

    # Remove last component
    def componentView_RemoveLastComponent(self):
        if self.componentsWidgetsLayout:
            component = self.componentsWidgetsLayout.pop()
            self.components_layout.removeWidget(component)
            component.deleteLater()

    # Update component view
    def componentView_UpdateComponentList(self):
        # Get the target number of components from the spinbox
        target_count = self.components_spinbox.value()
        # Get the current number of components in the list
        current_count = len(self.componentsClassList)

        # Adjust the number of components to match the target count
        while current_count != target_count:
            if target_count > current_count:
                # If more components are needed, create a new empty component
                newComponent = componentClass.ComponentClass("", "")

                # Add the new component to the components list
                self.componentsClassList.append(newComponent)

                # Create the visual representation of the component
                componentWidget = self.componenteView_createIndividualComponentView(newComponent)

                # Add the new component widget to the view
                self.componentView_AddComponent(componentWidget)

            elif target_count < current_count:
                # If fewer components are needed, remove the last component from the list
                self.componentsClassList.pop()

                # Remove the last component widget from the view
                self.componentView_RemoveLastComponent()

            # Update the current count after adding or removing a component
            current_count = len(self.componentsClassList)


    # Update quantities field
    def componentView_UpdateTotalQuantitiesField(self):
        boardsNumber = self.board_number_spinbox.value()
        for component in self.componentsWidgetsLayout:
            units_edit = component.findChild(QLineEdit, "OBJECT_NAME_UNIT")
            total_unit_edit = component.findChild(QLabel, "OBJECT_NAME_TOTAL_UNITS")
            total_unit_edit.setText(str(boardsNumber * int(units_edit.text())))

    # Read components from componentView
    def componentView_GetDataFromComponentView(self):
        self.componentsClassList = []

        # Iterate components view
        for component in self.componentsWidgetsLayout:

            # Get main data
            part_number_edit = component.findChild(QLineEdit, "OBJECT_NAME_PART_NUMBER")
            manufacturer_edit = component.findChild(QLineEdit, "OBJECT_NAME_MANUFACTURER")
            units_edit = component.findChild(QLineEdit, "OBJECT_NAME_UNIT")
            total_unit_edit = component.findChild(QLabel, "OBJECT_NAME_TOTAL_UNITS")

            # Create component
            componente = componentClass.ComponentClass(part_number_edit.text(), manufacturer_edit.text(),
                                                       units_edit.text(), total_unit_edit.text())

            #Iterate markets
            for row, market in enumerate(componentClass.markets, start=1):
                dataMarket = {}

                # Iterate attributes
                for idx, attribute in enumerate(componentClass.attributes):
                    if componentClass.attributes_visible[idx] == 0:
                        continue
                    # Generar un nombre único combinando market y attribute
                    object_name = f"{market}_{attribute.replace(' ', '_')}"
                    data = component.findChild(QLineEdit, object_name)
                    dataMarket[attribute] = data.text()
                componente.add_market_info(market, dataMarket)

            self.componentsClassList.append(componente)


    def handle_ok(self):
        # Get values from lists
        part_numbers = self.part_number_text.toPlainText().strip()
        manufacturers = self.manufacturer_text.toPlainText().strip()
        quantities = self.quantity_text.toPlainText().strip()

        # Split values by (CRLF)
        part_numbers_list = part_numbers.splitlines()
        manufacturers_list = manufacturers.splitlines()
        quantities_list = quantities.splitlines()

        # Verification
        if len(part_numbers_list) == len(manufacturers_list) == len(quantities_list):
            # Establece el proyecto como vacio
            self.empty_project()

            # Print each component as 'part_number manufacturer quantity'
            for part_number, manufacturer, quantity in zip(part_numbers_list, manufacturers_list, quantities_list):

                try:
                    quantity =  int(float(quantity))
                except ValueError:
                    quantity = 0

                # Create component
                componente = componentClass.ComponentClass(part_number,manufacturer,
                                                           quantity, 0)
                # Add component to list
                self.componentsClassList.append(componente)

                # Create and update component view
                componentWidget = self.componenteView_createIndividualComponentView(componente)
                self.componentView_AddComponent(componentWidget)
        else:
            QMessageBox.warning(self, "New project", "Error: Las listas de Part Numbers, Manufacturers y Quantities no tienen el mismo número de elementos.")

        # Close emergent view
        self.components_spinbox.setValue(len(self.componentsClassList))
        self.componentView_UpdateComponentList()
        self.dialog.reject()

    # Empty the project
    def empty_project(self):
        self.components_spinbox.setValue(0)

        # Iterate through the elements of the list
        for i in range(len(self.componentsClassList)):
            self.componentView_RemoveLastComponent()
        self.componentsClassList.clear()
        self.componentView_UpdateComponentList()

    # ======================================================
    #                 BUTTON ACTION METHODS
    # ======================================================

    # New project button
    def clickButtonAction_newProject(self):
        # Crear la ventana emergente (popup)
        self.dialog = QDialog(self)
        self.dialog.setWindowTitle("New project")

        # Layout principal
        layout = QVBoxLayout()

        # Crear los campos de texto para los part numbers, fabricantes y unidades
        self.part_number_text = QTextEdit()
        self.part_number_text.setPlaceholderText("Paste the Part Numbers here (one per line)")

        self.manufacturer_text = QTextEdit()
        self.manufacturer_text.setPlaceholderText("Paste the manufacturers here (one per line)")

        self.quantity_text = QTextEdit()
        self.quantity_text.setPlaceholderText("Paste the quantities here (one per line)")

        # Añadir los campos al layout
        layout.addWidget(QLabel("Part Numbers:"))
        layout.addWidget(self.part_number_text)

        layout.addWidget(QLabel("Manufacturers:"))
        layout.addWidget(self.manufacturer_text)

        layout.addWidget(QLabel("Quantity:"))
        layout.addWidget(self.quantity_text)

        # Botón OK y Cancelar
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

        # Conectar el botón OK a la función que maneja la acción
        button_box.accepted.connect(self.handle_ok)
        button_box.rejected.connect(self.dialog.reject)

        layout.addWidget(button_box)

        # Establecer el layout de la ventana emergente
        self.dialog.setLayout(layout)

        # Mostrar la ventana emergente
        self.dialog.exec_()


    # Save project button
    def clickButtonAction_saveProject(self):
        # Recupera los datos desde la vista
        self.componentView_GetDataFromComponentView()

        # Lista para almacenar los datos en formato JSON
        components_data = []

        # Agrega cada componente al arreglo con su formato JSON
        for component in self.componentsClassList:
            component_json = component.to_json()  # Supone que devuelve un diccionario
            components_data.append(component_json)

        # Permite al usuario seleccionar la ruta para guardar el archivo
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar proyecto",
            "",
            "Archivos JSON (*.json);;Todos los archivos (*)",
            options=options
        )

        if file_path:
            try:
                # Serialize the data to JSON and save it to the selected file
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump(components_data, file, indent=4, ensure_ascii=False)
                QMessageBox.information(self, "Save Project", "The project was saved successfully.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save the project:\n{e}")
        else:
            QMessageBox.warning(self, "Save Project", "No path was selected to save the project.")

    def clickButtonAction_openProject(self):

        # Crea un nuevo proyecto vacio
        self.empty_project()

        # Permite al usuario seleccionar un archivo JSON para abrir
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open project",
            "",
            "Archivos JSON (*.json);;Todos los archivos (*)",
            options=options
        )

        if file_path:
            try:
                # Abre el archivo JSON y carga los datos
                with open(file_path, 'r', encoding='utf-8') as file:
                    components_data = json.load(file)

                # Verifica que el archivo contiene una lista de componentes
                if not isinstance(components_data, list):
                    raise ValueError(
                        "wrong json data")

                # Limpia la lista actual de componentes
                self.componentsClassList.clear()
                # self.clear_component_views()

                # Crea instancias de ComponentClass para cada componente en los datos
                for component_data in components_data:
                    new_component = componentClass.ComponentClass(
                        part_number=component_data.get("part_number", ""),
                        manufacturer=component_data.get("manufacturer", ""),
                        units_per_board=component_data.get("units_per_board", 0),
                        total_units_per_board=component_data.get("total_units_per_board", 0)
                    )

                    # Agrega información de mercado al componente
                    market_info = component_data.get("market_info", {})
                    for market_name, market_data in market_info.items():
                        for attribute in componentClass.attributes:
                            if attribute not in market_data:
                                market_data[attribute] = ""

                        new_component.add_market_info(market_name, market_data)

                    # Añade el componente a la lista
                    self.componentsClassList.append(new_component)

                    # Crea y muestra la vista del componente
                    componentWidget = self.componenteView_createIndividualComponentView(new_component)
                    self.componentView_AddComponent(componentWidget)

                self.components_spinbox.setValue(len(components_data))
                QMessageBox.information(self, "Open Project", "The project was loaded successfully.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to open the project:\n{e}")
        else:
            QMessageBox.warning(self,"Open Project", "No file was selected to open.")

    def clickButtonAction_searchComponentsInShops(self):
        # Get config
        self.config = self.load_config()

        # Lee todos los componentes de la lista
        self.componentView_GetDataFromComponentView()

        # Get enabled markets
        markets_enabled = {
            market: self.config.getboolean("ENABLED", market, fallback=False)
            for market in componentClass.markets
        }
        # Init serach component shop class
        searcher = searchComponentInShopClass.SearchComponentInShop()

        # Iterate through the markets to initialize them if they are enabled
        for index, market in enumerate(componentClass.markets):

            if market == "MOUSER" and markets_enabled["MOUSER"]:
                token = self.config.get("TOKENS", "mouser_token", fallback="")
                # Init Api
                searcher.init_mouser(token)

            elif market == "DIGIKEY" and markets_enabled["DIGIKEY"]:
                clientID = self.config.get("TOKENS", "digikey_token", fallback="")
                clientSecret = self.config.get("TOKENS_SECRET", "digikey_token_secret", fallback="")
                digikeyLanguage = self.config.get("LANGUAGE", "digikey_language", fallback="")
                digikeyLocalSite = self.config.get("LOCAL_SITE", "digikey_local_site", fallback="")
                digikeyCurrency = self.config.get("CURRENCY", "digikey_currency", fallback="")
                # Init Api
                searcher.init_digikey(clientID, clientSecret, digikeyCurrency, digikeyLanguage, digikeyLocalSite)

            elif market == "TME" and markets_enabled["TME"]:
                tmetoken = self.config.get("TOKENS", "tme_token", fallback="")
                tmeSecret = self.config.get("TOKENS_SECRET", "tme_token_secret", fallback="")
                tmeCurrency = self.config.get("CURRENCY", "tme_currency", fallback="")
                tmeLanguage = self.config.get("LANGUAGE", "tme_language", fallback="")
                tmeLocalSite =self.config.get("LOCAL_SITE", "tme_local_site", fallback="")
                # Init Api
                searcher.init_tme(tmetoken, tmeSecret, tmeLocalSite, tmeLanguage, tmeCurrency)

            # Comprueba el mercado de Element14
            elif market == "ELEMENT14" and markets_enabled["ELEMENT14"]:
                token = self.config.get("TOKENS", "element14_token", fallback="")
                element14_market = self.config.get("MARKETS", "element14_market", fallback="")
                # Init Api
                searcher.init_element14(token, element14_market)

        # Iterate componts
        for component in self.componentsClassList:

            # Iterate markets
            for index, market in enumerate(componentClass.markets):

                # Search components
                partNumber = component.part_number
                quantity_value = int(component.total_units_per_board)
                partNumberResult = searcher.search_part_number(market, partNumber, quantity_value)
                #print(partNumberResult)

                message = "NotFound"
                info_alert = 2 #No Alert

                if partNumberResult and len(partNumberResult) == 1:
                    # Recupera la informacion
                    price_info = searcher.get_price_per_quantity(market, partNumberResult[0], quantity_value)

                    message = "unknown"
                    # Comprueba si se ha recomendado más cantidades
                    if ((partNumberResult[0]["availability"] > 0) & (quantity_value > partNumberResult[0]["availability"])):
                        message = "Insufficient stock."
                        info_alert = 1
                    elif partNumberResult[0]["availability"] > 0:
                        message = "Found."
                        info_alert = 0
                    elif (partNumberResult[0]["availability"] == 0) & (partNumberResult[0]["lifeCycleStatus"] == 'STOCKED'):
                        message = "Check on web"
                        info_alert = 2
                    elif partNumberResult[0]["availability"] == 0:
                        message = partNumberResult[0]["lifeCycleStatus"]
                        info_alert = 2
                    elif (price_info["recommended_units"] > quantity_value):
                        message = price_info["message"]

                    market_info = component.create_market_info(partNumberResult[0]["availability"],
                                                               price_info["recommended_price"],
                                                               price_info["total_price"],
                                                               price_info["recommended_units"], message, info_alert)
                else:
                    market_info = component.create_market_info("", "", "", "", "Not found.", info_alert)

                if markets_enabled[market] == True:
                    component.add_market_info(market, market_info)

        # Remove components views
        self.componentView_RemoveAllComponents()

        # Add again components
        for component in self.componentsClassList:
            componentWidget = self.componenteView_createIndividualComponentView(component)
            self.componentView_AddComponent(componentWidget)

    def clickButtonAction_generateOptimizedBom(self):
        file_dialog = QFileDialog()
        file_dialog.setAcceptMode(QFileDialog.AcceptSave)
        file_dialog.setDefaultSuffix(".xlsx")
        file_dialog.setNameFilters(["Excel Files (*.xlsx)"])

        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
        else:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("No file selected. Operation cancelled.")
            msg_box.setWindowTitle("Warning")
            msg_box.exec_()
            return

        # Create workbook
        wb = Workbook()

        # Get enabled markets
        markets_enabled = {
            market: self.config.getboolean("ENABLED", market, fallback=False)
            for market in componentClass.markets
        }

        not_found_components = []

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Iterate components
        for component in self.componentsClassList:
            best_market = None
            best_price = float('inf')
            best_units = 0
            best_total_price = 0
            best_stock = 0

            # Iterate enabled markets
            for market_name, market_data in component.market_info.items():
                if markets_enabled[market_name]:
                    stock = get_float_value(market_data, "STOCK")
                    total_price = get_float_value(market_data, "RECOMMENDED TOTAL PRICE")
                    recommended_units = get_float_value(market_data, "RECOMMENDED QUANTITY")

                    # Check stock
                    if int(component.total_units_per_board) <= stock:
                        if total_price < best_price:
                            best_market = market_name
                            best_price = total_price
                            best_units = recommended_units
                            best_total_price = total_price
                            best_stock = stock

            if best_market:
                # Add market info
                if best_market not in wb.sheetnames:
                    ws = wb.create_sheet(title=best_market)
                    ws.append(["Component Number", "Part Number", "Stock", "Quantity", "Total Price"])
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border_style
                else:
                    ws = wb[best_market]

                component_number = len(ws["A"])  # Contar las filas existentes para determinar el número del componente
                row = [component_number, component.part_number,best_stock,  best_units, best_total_price]
                ws.append(row)
                for cell in ws[len(ws["A"])]:
                    cell.border = border_style
            else:
                not_found_components.append(component.part_number)

        # Not found components view
        if not_found_components:
            ws_not_found = wb.create_sheet(title="NOT FOUND")
            ws_not_found.append(["Part Number"])
            for cell in ws_not_found[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_style
            for part_number in not_found_components:
                ws_not_found.append([part_number])
                for cell in ws_not_found[len(ws_not_found["A"])]:
                    cell.border = border_style

        # Fix format workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name != "NOT FOUND":
                total_price_col = [row[4].value for row in ws.iter_rows(min_row=2, max_col=5) if
                                   row[4].value is not None]
                total_sum = sum(total_price_col)
                ws.append([])
                ws.append(["", "", "Total", total_sum])
                for cell in ws[len(ws["A"])]:
                    cell.font = Font(bold=True)
                    cell.border = border_style
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15

        # Save workbook
        wb.save(file_path)

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setText("File generated")
        msg_box.setWindowTitle("Info")
        msg_box.exec_()


    def clickButtonAction_openConfiguration(self):
        dialog = ConfigurationWindow()
        dialog.exec_()


if __name__ == "__main__":
    #Init application
    app = QApplication(sys.argv)

    #Create main Window
    window = MainWindow()
    window.show()

    sys.exit(app.exec_())
